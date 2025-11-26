import streamlit as st
import psycopg2
import pandas as pd
import bcrypt

# =====================================
# CONFIGURACIÓN INICIAL
# =====================================
st.set_page_config(page_title="Sistema de Asistencias - UdeC", page_icon="🎓")
st.title("🎓 Sistema de Control de Asistencias")

# =====================================
# CONTROL DE SESIÓN
# =====================================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
    st.session_state.rol = None
    st.session_state.user_id = None
    st.session_state.nombre = None


# =====================================
# FUNCIÓN DE CONEXIÓN
# =====================================
def get_connection():
    return psycopg2.connect(st.secrets["postgres"]["url"])


# =====================================
# VERIFICAR CONTRASEÑA
# =====================================
def verificar_password(password_ingresada, password_bd):
    if password_bd is None:
        return False
    try:
        return bcrypt.checkpw(password_ingresada.encode('utf-8'), password_bd.encode('utf-8'))
    except Exception:
        return password_ingresada == password_bd  # fallback si no está cifrada


# =====================================
# SI YA ESTÁ LOGUEADO
# =====================================
if st.session_state.logged_in:
    st.sidebar.write(f"👋 Hola, {st.session_state.nombre}")
    st.sidebar.write(f"Rol: {st.session_state.rol}")

    if st.sidebar.button("Cerrar sesión"):
        st.session_state.logged_in = False
        st.session_state.rol = None
        st.session_state.user_id = None
        st.session_state.nombre = None
        st.rerun()

    rol = st.session_state.rol
    no_cuenta = st.session_state.user_id

    try:
        conn = get_connection()
        cursor = conn.cursor()

        if rol == "Administrador":
            st.subheader("⚙️ Panel de Administración")

            opcion = st.sidebar.selectbox(
                "Selecciona una acción",
                ["Ver alumnos", "Ver profesores", "Ver materias", "Ver clases", "Ver alumnos/clase", "Agregar registros"]
            )

            if opcion == "Ver alumnos":
                cursor.execute("SELECT * FROM alumnos ORDER BY no_cuenta")
                st.dataframe(pd.DataFrame(cursor.fetchall(), columns=["No. Cuenta", "Nombre", "Carrera", "Grado", "Grupo", "Edad", "Correo", "Password"]))

            elif opcion == "Ver profesores":
                cursor.execute("SELECT * FROM profesores ORDER BY no_cuenta")
                st.dataframe(pd.DataFrame(cursor.fetchall(), columns=["No. Cuenta", "Nombre", "Facultad", "Carrera", "Correo", "Password"]))

            elif opcion == "Ver materias":
                cursor.execute("SELECT * FROM materias ORDER BY id_materia")
                st.dataframe(pd.DataFrame(cursor.fetchall(), columns=["ID Materia", "Nombre", "Carrera", "Grado", "Creditos"]))

            elif opcion == "Ver clases":
                cursor.execute("""
                    SELECT c.id_clase, m.nombre AS materia, c.grupo, p.nombre AS profesor
                    FROM clases c
                    JOIN materias m ON c.id_materia = m.id_materia
                    JOIN profesores p ON c.no_cuenta_maestro = p.no_cuenta
                    ORDER BY c.id_clase
                """)
                st.dataframe(pd.DataFrame(cursor.fetchall(), columns=["ID Clase", "Materia", "Grupo", "Profesor"]))

            elif opcion == "Ver alumnos/clase":
                clase_id = st.number_input("ID de clase", min_value=1)
                if st.button("Ver alumnos"):
                    cursor.execute("""
                        SELECT a.no_cuenta, a.nombre, asis.estado, asis.fecha
                        FROM asistencias asis
                        JOIN alumnos a ON a.no_cuenta = asis.no_cuenta_alumno
                        WHERE asis.id_clase = %s
                    """, (clase_id,))
                    df = pd.DataFrame(cursor.fetchall(), columns=["No. Cuenta", "Alumno", "Estado", "Fecha"])
                    st.dataframe(df)

            elif opcion == "Agregar registros":
                tabla = st.selectbox("¿A qué tabla deseas agregar?", ["alumnos", "profesores", "materias", "clases", "alumnos_clases"])
                if tabla == "alumnos":
                    with st.form("agregar_alumno"):
                        no_cuenta_nuevo = st.number_input("Número de cuenta", min_value=10000000, max_value=99999999)
                        nombre = st.text_input("Nombre completo")
                        carrera = st.text_input("Carrera")
                        grado = st.number_input("Grado", min_value=1, max_value=10)
                        grupo = st.text_input("Grupo", max_chars=1)
                        edad = st.number_input("Edad", min_value=15, max_value=100)
                        correo = st.text_input("Correo")
                        password = st.text_input("Contraseña")
                        enviar = st.form_submit_button("Guardar")
                        if enviar:
                            cursor.execute("""
                                INSERT INTO alumnos (no_cuenta, nombre, carrera, grado, grupo, edad, correo, password)
                                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                            """, (no_cuenta_nuevo, nombre, carrera, grado, grupo, edad, correo, password))
                            conn.commit()
                            st.success("✅ Alumno agregado correctamente.")
                if tabla == "profesores":
                    with st.form("agregar_profesor"):
                        no_cuenta_nuevo = st.number_input("Número de cuenta", min_value=10000000, max_value=99999999)
                        nombre = st.text_input("Nombre completo")
                        facultad = st.text_input("Facultad")
                        carrera = st.text_input("Carrera")
                        correo = st.text_input("Correo")
                        password = st.text_input("Contraseña")
                        enviar = st.form_submit_button("Guardar")
                        if enviar:
                            cursor.execute("""
                                INSERT INTO profesores (no_cuenta, nombre, facultad, carrera, correo, password)
                                VALUES (%s,%s,%s,%s,%s,%s)
                            """, (no_cuenta_nuevo, nombre, facultad, carrera, correo, password))
                            conn.commit()
                            st.success("✅ Profesor agregado correctamente.")
                if tabla == "materias":
                    with st.form("agregar_materia"):
                        id_materia = st.text_input("ID de materia", max_chars=8)
                        nombre = st.text_input("Nombre de la materia")
                        carrera = st.text_input("Carrera")
                        grado = st.number_input("Grado", min_value=1, max_value=10)
                        creditos = st.number_input("Créditos", min_value=1, max_value=10)
                        enviar = st.form_submit_button("Guardar")
                        if enviar:
                            cursor.execute("""
                                INSERT INTO materias (id_materia, nombre, carrera, grado, creditos)
                                VALUES (%s,%s,%s,%s,%s)
                            """, (id_materia, nombre, carrera, grado, creditos))
                            conn.commit()
                            st.success("✅ Materia agregada correctamente.")
                if tabla == "clases":
                    with st.form("agregar_clase"):
                        no_cuenta_maestro = st.number_input("Número de cuenta del profesor", min_value=10000000, max_value=99999999)
                        id_materia = st.text_input("ID de materia", max_chars=8)
                        grupo = st.text_input("Grupo", max_chars=1)
                        enviar = st.form_submit_button("Guardar")
                        if enviar:
                            cursor.execute("""
                                INSERT INTO clases (no_cuenta_maestro, id_materia, grupo)
                                VALUES (%s,%s,%s)
                            """, (no_cuenta_maestro, id_materia, grupo))
                            conn.commit()
                            st.success("✅ Clase agregada correctamente.")
                if tabla == "alumnos_clases":
                    with st.form("agregar_alumno_clase"):
                        no_cuenta_alumno = st.number_input("Número de cuenta del alumno", min_value=10000000, max_value=99999999)
                        id_clase = st.number_input("ID de clase", min_value=1)
                        enviar = st.form_submit_button("Guardar")
                        if enviar:
                            cursor.execute("""
                                INSERT INTO alumnos_clases (no_cuenta_alumno, id_clase)
                                VALUES (%s,%s)
                            """, (no_cuenta_alumno, id_clase))
                            conn.commit()
                            st.success("✅ Alumno asignado a la clase correctamente.")

        elif rol == "Profesor":
            st.subheader("📚 Tus clases")
            query = """
                SELECT c.id_clase, m.nombre AS materia, c.grupo
                FROM clases c
                JOIN materias m ON c.id_materia = m.id_materia
                WHERE c.no_cuenta_maestro = %s
            """
            cursor.execute(query, (no_cuenta,))
            clases = pd.DataFrame(cursor.fetchall(), columns=["ID Clase", "Materia", "Grupo"])
            st.dataframe(clases)

            st.subheader("🧾 Registrar asistencias (modo detallado)")

            clase_id = st.number_input("ID de clase", min_value=1)

            # Control para no recargar los alumnos
            if "lista_grupo" not in st.session_state:
                st.session_state.lista_grupo = False

            if st.button("Cargar alumnos"):
                st.session_state.lista_grupo = True
                st.session_state.clase_id = clase_id
                st.session_state.pop("asistencias_df", None)
                st.rerun()

            if st.session_state.get("lista_grupo", False):
                clase_id = st.session_state.clase_id

                # Obtener nombre de la materia
                cursor.execute("""
                    SELECT m.nombre 
                    FROM clases c
                    JOIN materias m ON c.id_materia = m.id_materia
                    WHERE c.id_clase = %s
                """, (clase_id,))
                materia_nombre = cursor.fetchone()[0]

                # Obtener alumnos
                cursor.execute("""
                    SELECT a.no_cuenta, a.nombre
                    FROM alumnos_clases al_cl
                    JOIN alumnos a ON a.no_cuenta = al_cl.no_cuenta_alumno
                    WHERE al_cl.id_clase = %s
                """, (clase_id,))
                
                alumnos = cursor.fetchall()

                if not alumnos:
                    st.warning("⚠️ No hay alumnos en esta clase.")
                else:
                    st.write("### Marca la asistencia del grupo:")

                    # Crear dataframe temporal si no existe
                    if "asistencias_df" not in st.session_state:
                        df = pd.DataFrame(alumnos, columns=["NoCuenta", "Nombre"])
                        df["Asiste"] = True
                        df["Justificado"] = False
                        st.session_state.asistencias_df = df

                    df_editado = st.data_editor(
                        st.session_state.asistencias_df,
                        num_rows="fixed",
                        key="editor_asistencias",
                        column_config={
                            "Asiste": st.column_config.CheckboxColumn(
                                "Asiste", help="Marca si el alumno asistió", default=True),
                            "Justificado": st.column_config.CheckboxColumn(
                                "Justificado", help="Marca si la falta fue justificada", default=False)
                        },
                        hide_index=True
                    )

                    st.session_state.asistencias_df = df_editado

                    fecha = st.date_input("Fecha")
                    hora = st.time_input("Hora")

                    # -------------------------------
                    #   REGISTRO DE ASISTENCIAS
                    # -------------------------------
                    if st.button("Registrar asistencias del grupo"):
                        try:
                            for _, row in df_editado.iterrows():
                                no_cuenta_alumno = row["NoCuenta"]

                                if row["Asiste"]:
                                    estado = "Presente"
                                else:
                                    estado = "Justificado" if row["Justificado"] else "Ausente"

                                cursor.execute("""
                                    INSERT INTO asistencias (no_cuenta_alumno, id_clase, fecha, hora, estado)
                                    VALUES (%s, %s, %s, %s, %s)
                                """, (no_cuenta_alumno, clase_id, fecha, hora, estado))

                            conn.commit()
                            st.success("✅ Todas las asistencias fueron registradas correctamente.")
                        
                        except Exception as e:
                            st.error(f"❌ Error al registrar: {e}")


                    # ============================================================
                    #   EXPORTAR ASISTENCIAS A EXCEL (CON NOMBRE MATERIA + DOCENTE)
                    # ============================================================
                    st.write("### 📤 Exportar asistencias a Excel")

                    if st.button("Descargar Excel de asistencias"):
                        try:
                            cursor.execute("""
                                SELECT a.no_cuenta, al.nombre, a.estado, a.fecha, a.hora
                                FROM asistencias a
                                JOIN alumnos al ON al.no_cuenta = a.no_cuenta_alumno
                                WHERE a.id_clase = %s
                                ORDER BY a.fecha DESC, al.nombre
                            """, (clase_id,))

                            asist_df = pd.DataFrame(cursor.fetchall(), 
                                columns=["No. Cuenta", "Alumno", "Estado", "Fecha", "Hora"])

                            if asist_df.empty:
                                st.warning("⚠️ No hay asistencias registradas para exportar.")
                            else:
                                import io
                                from openpyxl import Workbook

                                # Crear excel
                                output = io.BytesIO()
                                wb = Workbook()
                                ws = wb.active

                                # ENCABEZADO
                                ws["A1"] = "Reporte de Asistencias"
                                ws["A2"] = f"Materia: {materia_nombre}"
                                ws["A3"] = f"Profesor: {st.session_state.nombre}"
                                ws["A5"] = "Listado de Asistencias"

                                # Insertar DataFrame
                                for col_idx, col_name in enumerate(asist_df.columns, start=1):
                                    ws.cell(row=7, column=col_idx, value=col_name)

                                for row_idx, row in asist_df.iterrows():
                                    for col_idx, value in enumerate(row, start=1):
                                        ws.cell(row=row_idx + 8, column=col_idx, value=value)

                                wb.save(output)
                                excel_data = output.getvalue()

                                st.download_button(
                                    label="📥 Descargar archivo Excel",
                                    data=excel_data,
                                    file_name=f"Asistencias_{materia_nombre}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                )

                        except Exception as e:
                            st.error(f"❌ Error al exportar: {e}")

        elif rol == "Alumno":
            st.subheader("📘 Tus asistencias")
            query = """
                SELECT m.nombre AS materia, a.estado, a.fecha, a.hora
                FROM asistencias a
                JOIN clases c ON a.id_clase = c.id_clase
                JOIN materias m ON c.id_materia = m.id_materia
                WHERE a.no_cuenta_alumno = %s
            """
            cursor.execute(query, (no_cuenta,))
            asistencias = pd.DataFrame(cursor.fetchall(), columns=["Materia", "Estado", "Fecha", "Hora"])
            st.dataframe(asistencias)

    except Exception as e:
        st.error(f"Error al conectar: {e}")


# =====================================
# SI NO ESTÁ LOGUEADO (LOGIN)
# =====================================
else:
    rol = st.selectbox("Selecciona tu rol", ["Alumno", "Profesor", "Administrador"])
    no_cuenta = st.text_input("Número de cuenta")
    password = st.text_input("Contraseña", type="password")

    if st.button("Iniciar sesión"):
        try:
            conn = get_connection()
            cursor = conn.cursor()

            if rol == "Administrador":
                cursor.execute("SELECT nombre, password FROM administradores WHERE usuario = %s", (no_cuenta,))
            else:
                tabla = "alumnos" if rol == "Alumno" else "profesores"
                cursor.execute(f"SELECT nombre, password FROM {tabla} WHERE no_cuenta = %s", (no_cuenta,))
            user = cursor.fetchone()

            if not user:
                st.error("❌ Usuario o número de cuenta no encontrado.")
            elif verificar_password(password, user[1]):
                st.session_state.logged_in = True
                st.session_state.rol = rol
                st.session_state.user_id = no_cuenta
                st.session_state.nombre = user[0]
                st.rerun()
            else:
                st.error("❌ Contraseña incorrecta.")
        except Exception as e:
            st.error(f"Error al conectar: {e}")
